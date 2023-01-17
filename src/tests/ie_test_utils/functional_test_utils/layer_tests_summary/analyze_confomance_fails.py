# Copyright (C) 2022-2023 Intel Corporation
# SPDX-License-Identifier: Apache-2.0

import os
import re
import shutil
import zipfile
import argparse
import openpyxl

from utils import utils


LOGS_ZIP_NAME = 'logs.zip'
JOB_PREFIX_LINUX = 'ie-tests-linux_'
NEW_LOG_DIR = 'ie_logs'

SW_PLUGINS = {'HETERO': '1', 'AUTO': '2', 'BATCH': '3', 'MULTI': '4'}

logger = utils.get_logger('AnalyzerConformanceLog')


class AnalyzerConformanceLog:
    def __init__(self):
        self.analyzed_hw_devices = set()
        self.local_log_dir = os.path.join(os.getcwd())
        self.logs_info = {}


    def setup_log_local_dir(self, user_log_dir):
        if user_log_dir:
            if os.path.exists(user_log_dir):
                self.local_log_dir = user_log_dir
            else:
                logger.error(f'Local log dir {user_log_dir} is not exists')
                return 1


    def process_remote_log(self, remote_log_path, job_list):
        if remote_log_path and job_list:
            logger.info("Downloading and unziping remote log")
            # if something will be downloaded, create new dir for these purpose
            self.local_log_dir = os.path.join(self.local_log_dir, NEW_LOG_DIR)

            if 'linux' not in remote_log_path:
                logger.error('Sorry, I can analyze just linux results now')
                return 1

            for job_number in job_list:
                job_name = f'{JOB_PREFIX_LINUX}{job_number}'
                source_path = os.path.join(remote_log_path, job_name, LOGS_ZIP_NAME)
                dest_path = os.path.join(self.local_log_dir, job_number)
                logger.info(f'Copying {source_path} to {dest_path}')
                try:
                    os.makedirs(dest_path, exist_ok=True)
                    shutil.copy2(source_path, dest_path)
                except Exception as e:
                    logger.error(f'FAIL {e}')
                else:
                    logger.info(f'DONE')

                logger.info(f'Unziping {dest_path}')
                extended_path = "\\\\?\\%s" % os.path.join(dest_path, LOGS_ZIP_NAME)
                with zipfile.ZipFile(extended_path, "r") as zipObj:
                    for name in zipObj.namelist():
                        if re.match(f'test/apiconformancetests_(\w*)_(dlb|omz)-\d.log', name):
                            zipObj.extract(name, dest_path)
                            logger.error(f'OK')


    def collect_tests_result(self, exclude_from_log):
        logger.info(f'Collecting results')
        for root, dirs, files in os.walk(self.local_log_dir, topdown=False):
            for file_name in files:
                file_name_match = re.match(f'apiconformancetests_(\w*)_(dlb|omz)-\d.log', file_name)
                if "test" in root and file_name_match:
                    extended_path = "\\\\?\\%s" % os.path.join(root, file_name)
                    try:
                        lines = ''
                        with open(extended_path, encoding="utf8") as f:
                            lines = f.readlines()

                        device = file_name_match.group(1)
                        self.analyzed_hw_devices.add(device)

                        test_name = None
                        test_number = None
                        in_run_stage = False
                        error_msg = ''
                        test_info_by_device = None

                        for line in lines:
                            if utils.RUN in line:
                                in_run_stage = True
                                error_msg = ''
                                # if run stage exists, it is because gtest decided to show log as test fails
                                test_info_by_device['pass'] = False
                                continue

                            # it is result, we got to the end of run stage
                            if utils.TEST_STATUS['failed'] in line:
                                in_run_stage = False
                                if error_msg:
                                    test_info_by_device['err_info'] = error_msg
                                    error_msg = None
                                continue
                            
                            # collect error message in run stage
                            if in_run_stage:
                                # remove date
                                line = re.sub('\[\d*-\d*-\d* \d*:\d*:\d*,\d*\] \[\d*\] .* INFO: ', '', line)
                                line = line.strip('\n')
                                line = line.strip(' ')
                                if exclude_from_log:
                                    error_msg = error_msg.replace(exclude_from_log, '')
                                if line and 'MEM_USAGE' not in line:
                                    error_msg += ' ' + line + ' '

                            test_match = re.search(r'(\[\d*/\d*\]) (.*) \(\d* ms\)', line)
                            if test_match:
                                in_run_stage = False
                                # analyzed failed test without RUN stage, it can be crashed tests
                                if error_msg and test_info_by_device:
                                    error_msg = error_msg.replace(test_info_by_device['name'], '')
                                    error_msg = re.sub('\[\d*\/\d*\]', '', error_msg)
                                    error_msg = re.sub('\(\d* ms\)', '', error_msg)
                                    if exclude_from_log:
                                        error_msg = error_msg.replace(exclude_from_log, '')
                                    test_info_by_device['err_info'] = error_msg

                                error_msg = None

                                # start analyze new test
                                if test_number != test_match.group(1):
                                    test_number = test_match.group(1)
                                    test_name = test_match.group(2)
                                    test_name = re.sub(device.upper(), '[HWDevice]', test_name)
                                    test_group = test_name.split('/')[0] if len(test_name.split('/')) > 1 else test_name.split('.')[0]
                                    # cover case ov_infer_request_1/2/...
                                    if 'ov_infer_request_' in test_group:
                                        test_group = 'ov_infer_request'
                                    # cover case conformance_query_model_[ops]
                                    if 'conformance_query_model_' in test_group:
                                        test_group = 'conformance_query_model'

                                    self.logs_info.setdefault(test_group, {})
                                    self.logs_info[test_group].setdefault(test_name, {})
                                    self.logs_info[test_group][test_name][device] = {'name': test_match.group(2), 'pass': True, 'err_info': ''}
                                    test_info_by_device = self.logs_info[test_group][test_name][device]

                    except Exception as e:
                        logger.error(f'Analyzing of {file_name} FAIL: {e}')

    def create_exel(self, expected_devices, exclude_sw_plugins):
        logger.info(f'Creating exel file with results')

        if not expected_devices:
            expected_devices = self.analyzed_hw_devices

        workbook = openpyxl.workbook.Workbook()

        fail_font = openpyxl.styles.Font(color="9C0006")
        fail_fill = openpyxl.styles.PatternFill(start_color="FF8B8B", end_color="FF8B8B", fill_type = "solid")

        pass_font = openpyxl.styles.Font(color="006100")
        pass_fill = openpyxl.styles.PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type = "solid")

        not_run_font = openpyxl.styles.Font(color="000000")
        not_run_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type = "solid")


        for group_name, group_info in self.logs_info.items():
            sheet_name = group_name if len(group_name) < 30 else group_name[0:30]
            worksheet = workbook.create_sheet(sheet_name)

            worksheet['A1'] = 'Test'
            worksheet['B1'] = 'Device'
            worksheet['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')
            worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=(len(expected_devices)*2 + 1))

            column = 2
            for dev in expected_devices:
                worksheet.cell(2, column, dev.upper())
                worksheet.cell(2, column).alignment = openpyxl.styles.Alignment(horizontal='center')
                worksheet.cell(2, column).font = openpyxl.styles.Font(bold=True)
                worksheet.merge_cells(start_row=2, start_column=column, end_row=2, end_column=(column + 1))
                worksheet.cell(3, column, 'Status')
                column += 1
                worksheet.cell(3, column, 'Error')
                column += 1

            row = 4
            column = 1

            for test_name, devices_info in group_info.items():
                if exclude_sw_plugins and re.search(exclude_sw_plugins, test_name):
                    continue

                # if test pass on all devices, it is not need to analyze it here
                all_pass = all([dev.lower() in devices_info and devices_info[dev.lower()]['pass'] for dev in expected_devices])
                if all_pass:
                    continue

                worksheet.cell(row, column, test_name)
                column += 1
                for dev in expected_devices:
                    status = 'Not run'
                    error_msg = ''
                    font = not_run_font
                    fill = not_run_fill
                    if dev.lower() in devices_info:
                        if devices_info[dev.lower()]['pass']:
                            font = pass_font
                            fill = pass_fill
                            status = 'Pass'
                        else:
                            status = 'Fail'
                            error_msg = devices_info[dev.lower()]['err_info']
                            font = fail_font
                            fill = fail_fill

                    worksheet.cell(row, column, status)
                    worksheet.cell(row, column).fill = fill
                    worksheet.cell(row, column).font = font
                    worksheet
                    column += 1
                    worksheet.cell(row, column, error_msg)

                    column += 1
                column = 1
                row += 1

        for worksheet in workbook.worksheets:
            if worksheet.max_row <= 3:
                workbook.remove(worksheet)

        workbook.save(os.path.join(self.local_log_dir, 'apiConformanceFails.xlsx'))


if __name__ == '__main__':

    parser = argparse.ArgumentParser()

    parser.add_argument('--log_dir',
                        required=False,
                        type=str,
                        help='Setup log folder for local machine.\
It can be folder with already downloaded and unzipped logs or folder where will be downloaded logs from remote share.\
If it is not be setup, folder of running script will be used as local log dir.')
    parser.add_argument('--remote_log_path',
                        required=False, type=str,
                        help=f'Setup path to folder on share with jenkins log for needed platform.\
To download logs folder with name {NEW_LOG_DIR} will be created \
If it is not be setup, it will be investigated log_dir to find logs. Logs should be unzipped.')
    parser.add_argument('--job_list',
                        required=False,
                        nargs='*',
                        help='Setup list of job number. Logs of these jobs will be downloaded, unzipped and anazized \
Example: 6030 6035')
    parser.add_argument("--expected_devices",
                        required=False,
                        nargs="*",
                        help='Setup hw devices name, if it is not be setup, all founded deviced will be presented in final report\
Example: CPU TEMPLATE')
    parser.add_argument('--exclude_from_log',
                        type=str,
                        help='This arguments could be use to removing repeated pattern in log, for example long paths of files with code')
    parser.add_argument('--exclude_sw_plugins',
                        required=False,
                        nargs="*",
                        help='Setup sw plagins to exclude. If it is not be setup, all plugins will be in report.\
Example1 - exclude several: AUTO HETERO\
Example2 - exclude all from AUTO/HETERO/MULTI/BATCH: ALL')

    args = parser.parse_args()

    AnalyzerConformanceLog = AnalyzerConformanceLog()

    if args.log_dir:
        if AnalyzerConformanceLog.setup_log_local_dir(args.log_dir):
            exit(1)

    if AnalyzerConformanceLog.process_remote_log(args.remote_log_path, args.job_list):
        exit(1)

    AnalyzerConformanceLog.collect_tests_result(args.exclude_from_log)

    exclude_sw_plugins = None
    if args.exclude_sw_plugins:
        if 'ALL' in args.exclude_sw_plugins:
            exclude_sw_plugins = '(' + '|'.join(SW_PLUGINS.keys()) + '|\/[1234]$)'
        else:
            exclude_sw_plugins = '(' + '|'.join(args.exclude_sw_plugins)
            sw_plugins_numbers = ''.join([SW_PLUGINS[plugin.upper()] for plugin in args.exclude_sw_plugins if plugin.upper() in SW_PLUGINS])
            exclude_sw_plugins += '|\/[' +  + ']&)' if sw_plugins_numbers else ')'
    AnalyzerConformanceLog.create_exel(args.expected_devices, exclude_sw_plugins)

