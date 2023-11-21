# Copyright (C) 2022 Intel Corporation
# SPDX-License-Identifier: Apache-2.0

import os
import re
import sys
import shutil
import zipfile
import argparse
import openpyxl

from utils.conformance_utils import get_logger
from utils import constants

LOGS_ZIP_NAME = 'logs.zip'
NEW_LOG_DIR = 'ie_logs'

SW_PLUGINS = {'HETERO': '1', 'AUTO': '2', 'BATCH': '3', 'MULTI': '4'}

logger = get_logger('AnalyzerConformanceLog')


class AnalyzerConformanceLog:
    def __init__(self):
        self.analyzed_hw_devices = set()
        self.local_log_dir = os.path.join(os.getcwd())
        self.logs_info = {}
        self.output_path = os.path.join(os.getcwd())
        self.output_file_name = "apiConformanceFails.xlsx"

    @staticmethod
    def status_folder_exists(path_name):
        return any([(status in path_name) for status in constants.TEST_STATUS.keys()])

    def setup_log_local_dir(self, user_log_dir):
        if user_log_dir:
            if os.path.exists(user_log_dir):
                self.local_log_dir = user_log_dir
            else:
                logger.error(f'Local log dir {user_log_dir} is not exists')

    def setup_output_file(self, output_path, output_file_name):
        if output_path:
            if os.path.exists(output_path):
                self.output_path = output_path
            else:
                logger.error(f'Output file path {output_path} is not exists, directory {self.output_path} will be used')

        if output_file_name:
            self.output_file_name = output_file_name

    def process_remote_log(self, remote_log_path, job_list):
        if remote_log_path and job_list:
            logger.info("Downloading and unziping remote log")
            # if something will be downloaded, create new dir for these purpose
            self.local_log_dir = os.path.join(self.local_log_dir, NEW_LOG_DIR)

            for job_name in job_list:
                source_path = os.path.join(remote_log_path, job_name, LOGS_ZIP_NAME)
                dest_path = os.path.join(self.local_log_dir, job_name)
                logger.info(f'Copying {source_path} to {dest_path}')

                try:
                    os.makedirs(dest_path, exist_ok=True)
                    shutil.copy2(source_path, dest_path)
                except Exception as e:
                    logger.error(f'FAIL {e}')
                    return 1
                else:
                    logger.info(f'DONE')

                logger.info(f'Unziping {dest_path}')
                extended_path = os.path.join(dest_path, LOGS_ZIP_NAME)
                if sys.platform.startswith('win'):
                    extended_path = "\\\\?\\%s" % os.path.join(dest_path, LOGS_ZIP_NAME)
                with zipfile.ZipFile(extended_path, "r") as zipObj:
                    for name in zipObj.namelist():
                        if 'apiconformancetests' in name and 'logs' in name and self.status_folder_exists(name):
                            zipObj.extract(name, dest_path)
                logger.info(f'DONE')

    def get_real_device(self, device):
        real_device = device.upper()
        if device.upper() == 'DGPU':
            real_device = 'GPU'
        elif device.upper() == 'ARM':
            real_device = 'CPU'

        return real_device

    def collect_tests_result(self, exclude_from_log):
        logger.info(f'Collecting results')
        for root, dirs, files in os.walk(self.local_log_dir, topdown=False):
            if not 'apiconformancetests' in root or not 'logs' in root or\
               not self.status_folder_exists(root):
                continue

            for file_name in files:
                extended_path = os.path.join(root, file_name)
                if sys.platform.startswith('win'):
                    extended_path = "\\\\?\\%s" % os.path.join(root, file_name)
                try:
                    lines = ''
                    with open(extended_path, encoding="utf8") as f:
                        lines = f.readlines()

                    path_components = os.path.normpath(extended_path).split(os.path.sep)
                    # examle: /cpu_conformance_dlb_apiconformancetests_nightly/logs/passed/[hash].log
                    retult_root_match = re.match(r'(\w*)_conformance_(dlb|omz)_apiconformancetests_.*', path_components[-4])
                    device = retult_root_match.group(1)
                    self.analyzed_hw_devices.add(device)

                    status = os.path.normpath(extended_path).split(os.path.sep)[-2]

                    test_name = None
                    error_msg = ''

                    for line in lines:
                        if constants.RUN in line:
                            test_name_match = re.match(f'.*\[\s*RUN\s*\] (.*)', line)
                            if test_name_match:
                                test_name = test_name_match.group(1)
                            continue

                        if line == '' or constants.MEM_USAGE in line or constants.REF_COEF in line or\
                           status == 'skipped' or status == 'passed' or constants.TEST_STATUS['failed'][0] in line:
                            continue

                        line = line.strip('\n')
                        line = line.strip(' ')
                        if exclude_from_log:
                            line = line.replace(exclude_from_log, '')
                        error_msg += ' ' + line + ' '

                    test_name = test_name or file_name
                    real_device = self.get_real_device(device)
                    test_name = re.sub(f'^(.+/.+\..+/.*?){real_device}(.*)$', r'\1[HWDevice]\2', test_name, count=1)
                    if real_device in SW_PLUGINS:
                        test_name = re.sub(f'(configItem|config)=.*$', '', test_name)

                    test_group = test_name.split('/')[0] if len(test_name.split('/')) > 1 else test_name.split('.')[0]
                    # cover case ov_infer_request_1/2/...
                    if 'ov_infer_request_' in test_group:
                        test_group = 'ov_infer_request'
                    # cover case conformance_query_model_[ops]
                    if 'conformance_query_model_' in test_group:
                        test_group = 'conformance_query_model'

                    self.logs_info.setdefault(test_group, {})
                    self.logs_info[test_group].setdefault(test_name, {})
                    self.logs_info[test_group][test_name][device] = {'name': test_name, 'status': status, 'err_info': error_msg}

                except Exception as e:
                    logger.error(f'Analyzing of {extended_path} FAIL: {e}')

    def create_exel(self, expected_devices):
        logger.info(f'Creating exel file with results')

        if not expected_devices:
            expected_devices = self.analyzed_hw_devices

        workbook = openpyxl.workbook.Workbook()

        fail_font = openpyxl.styles.Font(color="9C0006")
        fail_fill = openpyxl.styles.PatternFill(start_color="FF8B8B", end_color="FF8B8B", fill_type = "solid")

        pass_font = openpyxl.styles.Font(color="006100")
        pass_fill = openpyxl.styles.PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type = "solid")

        not_run_font = openpyxl.styles.Font(color="000000")
        not_run_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type = "solid")


        for group_name, group_info in self.logs_info.items():
            sheet_name = group_name if len(group_name) < 30 else group_name[0:30]
            worksheet = workbook.create_sheet(sheet_name)

            worksheet['A1'] = 'Test'
            worksheet['B1'] = 'Device'
            worksheet['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')
            worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=(len(expected_devices)*2 + 1))

            column = 2
            for dev in expected_devices:
                worksheet.cell(2, column, dev.upper())
                worksheet.cell(2, column).alignment = openpyxl.styles.Alignment(horizontal='center')
                worksheet.cell(2, column).font = openpyxl.styles.Font(bold=True)
                worksheet.merge_cells(start_row=2, start_column=column, end_row=2, end_column=(column + 1))
                worksheet.cell(3, column, 'Status')
                column += 1
                worksheet.cell(3, column, 'Error')
                column += 1

            row = 4
            column = 1

            for test_name, devices_info in group_info.items():
                # if test passed or skipped on all devices, it is not need to analyze it here
                all_pass = all([dev.lower() not in devices_info or (dev.lower() in devices_info and\
                                                                     (devices_info[dev.lower()]['status'] == 'passed' or\
                                                                      devices_info[dev.lower()]['status'] == 'skipped')) for dev in expected_devices])
                if all_pass:
                    continue

                worksheet.cell(row, column, test_name)
                column += 1
                for dev in expected_devices:
                    status = 'Not run'
                    error_msg = ''
                    font = not_run_font
                    fill = not_run_fill

                    real_dev = dev
                    if dev == "CPU_ARM":
                        real_dev = "arm"

                    if real_dev.lower() in devices_info:
                        if devices_info[real_dev.lower()]['status'] == 'passed':
                            font = pass_font
                            fill = pass_fill
                            status = 'Pass'
                        elif devices_info[real_dev.lower()]['status'] == 'failed' or devices_info[real_dev.lower()]['status'] == 'hanged' or\
                             devices_info[real_dev.lower()]['status'] == 'crashed' or devices_info[real_dev.lower()]['status'] == 'interapted':
                            status = 'Fail'
                            error_msg = devices_info[real_dev.lower()]['err_info']
                            font = fail_font
                            fill = fail_fill

                    worksheet.cell(row, column, status)
                    worksheet.cell(row, column).fill = fill
                    worksheet.cell(row, column).font = font
                    worksheet
                    column += 1
                    worksheet.cell(row, column, error_msg)

                    column += 1
                column = 1
                row += 1

        for worksheet in workbook.worksheets:
            if worksheet.max_row <= 3:
                workbook.remove(worksheet)

        workbook.save(os.path.join(self.output_path, self.output_file_name))


if __name__ == '__main__':

    parser = argparse.ArgumentParser()

    parser.add_argument('--log_dir',
                        required=False,
                        type=str,
                        help='Setup log folder for local machine.\
It can be folder with already downloaded and unzipped logs or folder where will be downloaded logs from remote share.\
If it is not be setup, folder of running script will be used as local log dir.')
    parser.add_argument('--remote_log_path',
                        required=False, type=str,
                        help=f'Setup path to folder on share with jenkins log for needed platform.\
To download logs folder with name {NEW_LOG_DIR} will be created \
If it is not be setup, it will be investigated log_dir to find logs. Logs should be unzipped.')
    parser.add_argument('--job_list',
                        required=False,
                        nargs='*',
                        help='Setup list of jobs. Logs of these jobs will be downloaded, unzipped and anazized \
Example: job_name1 job_name2')
    parser.add_argument("--expected_devices",
                        required=False,
                        nargs="*",
                        help='Setup hw devices name, if it is not be setup, all founded deviced will be presented in final report\
Example: CPU TEMPLATE')
    parser.add_argument("--output_path",
                        required=False,
                        type=str,
                        help='Setup path for output xlsx file')
    parser.add_argument("--output_file_name",
                        required=False,
                        type=str,
                        help='Setup name for output xlsx file')
    parser.add_argument('--exclude_from_log',
                        type=str,
                        help='This arguments could be use to removing repeated pattern in log, for example long paths of files with code')

    args = parser.parse_args()

    analyzerConformanceLog = AnalyzerConformanceLog()

    if args.log_dir:
        if analyzerConformanceLog.setup_log_local_dir(args.log_dir):
            exit(1)

    analyzerConformanceLog.setup_output_file(args.output_path, args.output_file_name)

    if analyzerConformanceLog.process_remote_log(args.remote_log_path, args.job_list):
        exit(1)

    analyzerConformanceLog.collect_tests_result(args.exclude_from_log)

    analyzerConformanceLog.create_exel(args.expected_devices)

