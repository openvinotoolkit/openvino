# Copyright (C) 2020-2022 Intel Corporation
# SPDX-License-Identifier: Apache-2.0

import os

import pytest
from addict import Dict
from openpyxl import Workbook, load_workbook

from openvino.tools.pot.app.run import optimize
from openvino.tools.pot.utils.logger import init_logger
from .utils.config import get_engine_config, merge_configs

init_logger(level='INFO')

TEST_MODELS_DEFAULT = [
    ('ssd512', 'caffe', 'FP32', {
        'performance': {'map': 0.9088},
        'mixed': {'map': 0.904}}),
    ('googlenet-v3', 'tf', 'FP32', {
        'performance': {'accuracy@top1': 0.7793, 'accuracy@top5': 0.9378},
        'mixed': {'accuracy@top1': 0.7793, 'accuracy@top5': 0.9378}}),
    ('squeezenet1.1', 'caffe', 'FP32', {
        'performance': {'accuracy@top1': 0.5772, 'accuracy@top5': 0.808},
        'mixed': {'accuracy@top1': 0.57706, 'accuracy@top5': 0.808}}),
    ('mobilenet-v1-1.0-224-tf', 'tf', 'FP32', {
        'performance': {'accuracy@top1': 0.70896, 'accuracy@top5': 0.89792},
        'mixed': {'accuracy@top1': 0.70922, 'accuracy@top5': 0.89806}}),
    ('mobilenet-v2-pytorch', 'pytorch', 'FP16', {
        'performance': {'accuracy@top1': 0.71552, 'accuracy@top5': 0.90222},
        'mixed': {'accuracy@top1': 0.71512, 'accuracy@top5': 0.90172}}),
    ('resnet-50-pytorch', 'pytorch', 'FP32', {
        'performance': {'accuracy@top1': 0.75936, 'accuracy@top5': 0.92854},
        'mixed': {'accuracy@top1': 0.75964, 'accuracy@top5': 0.92816}}),
    ('googlenet-v3-pytorch', 'pytorch', 'FP32', {
        'performance': {'accuracy@top1': 0.77562, 'accuracy@top5': 0.9363},
        'mixed': {'accuracy@top1': 0.77562, 'accuracy@top5': 0.9363}}),
    ('densenet-121', 'caffe', 'FP32', {
        'performance': {'accuracy@top1': 0.73908, 'accuracy@top5': 0.91728},
        'mixed': {'accuracy@top1': 0.7389, 'accuracy@top5': 0.91714}}),
    ('mobilenet-ssd', 'caffe', 'FP32', {
        'performance': {'map': 0.666},
        'mixed': {'map': 0.664}}),
    ('octave-resnet-26-0.25', 'mxnet', 'FP32', {
        'performance': {'accuracy@top1': 0.7581, 'accuracy@top5': 0.9256},
        'mixed': {'accuracy@top1': 0.759, 'accuracy@top5': 0.92466}}),
    ('ssd_mobilenet_v1_coco', 'tf', 'FP16', {
        'performance': {'coco_precision': 0.2312},
        'mixed': {'coco_precision': 0.2314}})
]

TEST_MODELS_ACC_AWARE = [
    ('efficientnet-b0-pytorch', 'pytorch', 'CPU', {'performance': {'accuracy@top1': 0.7663,
                                                                   'accuracy@top5': 0.9294}}),
    ('mobilenet-ssd', 'caffe', 'CPU', {'performance': {'map': 0.67}}),
    ('ssd512', 'caffe', 'CPU', {'performance': {'map': 0.7917}}),
    ('mobilenet-v1-0.25-128', 'tf', 'GNA', {'performance': {'accuracy@top1': 0.4133, 'accuracy@top5': 0.6626}})
]


def run_algo(config, model_name, model_framework, metrics, expected_result, tmp_path):
    result = optimize(config)
    metrics.update(result)
    write_results_to_xlsx(model_name, model_framework, result, expected_result, tmp_path)


@pytest.mark.parametrize('model_params', TEST_MODELS_DEFAULT,
                         ids=['{}_{}'.format(m[0], m[1]) for m in TEST_MODELS_DEFAULT])
def test_default_quantization(model_params, tmp_path, models, algorithm, preset):
    if algorithm != 'DefaultQuantization':
        pytest.skip()

    algorithm_config = Dict({
        'algorithms': [{
            'name': 'DefaultQuantization',
            'params': {
                'target_device': 'CPU',
                'stat_subset_size': 1000,
                'preset': preset
            }}]})

    model_name, model_framework, model_precision, expected_accuracy_dict = model_params
    run_quantization(models=models,
                     model_name=model_name,
                     model_framework=model_framework,
                     model_precision=model_precision,
                     algorithm_config=algorithm_config,
                     expected_accuracy=expected_accuracy_dict[preset],
                     tmp_path=tmp_path)


@pytest.mark.parametrize('model_params', TEST_MODELS_ACC_AWARE,
                         ids=['{}_{}'.format(m[0], m[1]) for m in TEST_MODELS_ACC_AWARE])
def test_accuracy_aware_quantization(model_params, tmp_path, models, algorithm, preset):
    if algorithm != 'AccuracyAwareQuantization':
        pytest.skip()

    model_name, model_framework, device, expected_accuracy_dict = model_params

    algorithm_config = Dict({
        'algorithms': [{
            'name': 'AccuracyAwareQuantization',
            'params': {
                'target_device': device,
                'base_algorithm': 'DefaultQuantization',
                'preset': preset,
                'stat_subset_size': 1000,
                'max_iter_num': 10,
                'maximal_drop': 0.004,
                'metric_subset_ratio': 0.5,
                'ranking_subset_size': 300,
                'drop_type': 'absolute'
            }}]})

    run_quantization(models=models,
                     model_name=model_name,
                     model_framework=model_framework,
                     algorithm_config=algorithm_config,
                     expected_accuracy=expected_accuracy_dict[preset],
                     tmp_path=tmp_path)


def run_quantization(models, model_name, model_framework, algorithm_config,
                     expected_accuracy, tmp_path, tolerance=0.0015, model_precision='FP32'):
    model = models.get(model_name, model_framework, tmp_path, model_precision)
    engine_config = get_engine_config(model_name)

    config = merge_configs(model.model_params, engine_config, algorithm_config)
    metrics = Dict()
    run_algo(config, model_name, model_framework, metrics, expected_accuracy, tmp_path)

    for metric, value in metrics.items():
        print('{}: {:.4f}'.format(metric, value))

    assert metrics == pytest.approx(expected_accuracy, abs=tolerance)


def write_results_to_xlsx(model_name, model_framework, metrics, expected_result, tmp_path):
    def adjust_columns_width():
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value) or '') + 1 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length

    save_path = (tmp_path.parent / 'results.xlsx').as_posix()
    if os.path.exists(save_path):
        workbook = load_workbook(save_path)
    else:
        workbook = Workbook()
    worksheet = workbook.active

    results = [model_name, model_framework]
    expected = ['', '']
    drop = ['', '']
    for name in metrics.keys():
        results.extend([name, round(metrics[name], 5)])
        expected.extend(['expected', round(expected_result[name], 5)])
        drop.extend(['drop', round(expected_result[name] - metrics[name], 5)])
    worksheet.append(results)
    worksheet.append(expected)
    worksheet.append(drop)

    adjust_columns_width()
    workbook.save(save_path)
